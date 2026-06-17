"""
Reporte Excel de oportunidades (+ envio por Telegram). Migrado de
main.py:generar_reporte_excel, esquema neutral (`url_tienda`/`precio_tienda`).
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .notifier import enviar_excel_telegram


def generar_reporte_excel(oportunidades, logger, enviar_telegram=True):
    if not oportunidades:
        logger.info("📄 No hay oportunidades para exportar a Excel")
        return None

    # Ordenar por prioridad (descendente)
    oportunidades = sorted(oportunidades, key=lambda op: op["prioridad"], reverse=True)

    os.makedirs("reports", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = f"reports/oportunidades_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Oportunidades"

    headers = [
        "Nombre", "Expansión", "Foil",
        "Precio Tienda (COP)", "Precio SCG (COP)",
        "Ganancia %", "Diferencia",
        "URL Tienda", "URL SCG", "Imagen",
    ]
    ws.append(headers)

    for op in oportunidades:
        ws.append([
            op["nombre"],
            op["expansion"],
            "Foil" if op["foil"] else "No Foil",
            op["precio_tienda"],
            op["precio_scg"],
            op["porcentaje"],
            op["diferencia"],
            "", "", "",  # URL Tienda / URL SCG / Imagen (se ponen como hyperlink)
        ])

        row = ws.max_row

        if op.get("url_tienda"):
            cell = ws[f"H{row}"]
            cell.value = "Abrir Tienda"
            cell.hyperlink = op["url_tienda"]
            cell.style = "Hyperlink"

        if op.get("url_scg"):
            cell = ws[f"I{row}"]
            cell.value = "Abrir SCG"
            cell.hyperlink = op["url_scg"]
            cell.style = "Hyperlink"

        if op.get("image_url"):
            cell = ws[f"J{row}"]
            cell.value = "Ver imagen"
            cell.hyperlink = op["image_url"]
            cell.style = "Hyperlink"

    # Auto ancho columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 45)

    wb.save(ruta)
    logger.info(f"📊 Reporte Excel generado: {ruta}")

    if enviar_telegram:
        try:
            enviar_excel_telegram(ruta, caption="📊 Reporte Excel de oportunidades detectadas")
            logger.info("📤 Reporte Excel enviado a Telegram correctamente")
        except Exception as e:
            logger.error(f"❌ Error enviando Excel a Telegram: {e}")

    return ruta
